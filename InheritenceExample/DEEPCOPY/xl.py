import openpyxl

# Open an existing .xlsx file
loc = 'demo.xlsx'
wb = openpyxl.load_workbook(loc)
sheet = wb.active  # Get the first sheet

# Extract data from the first row
row_index = 1  # openpyxl uses 1-based indexing
data = [sheet.cell(row=row_index, column=col).value for col in range(1, sheet.max_column + 1)]

# Create a new workbook
wb1 = openpyxl.Workbook()
sheet1 = wb1.active
sheet1.title = 'dummy'

for index, value in enumerate(data, start=1):
    sheet1.cell(row=1, column=index, value=value)

# Save to a new .xlsx file
wb1.save('output.xlsx')
